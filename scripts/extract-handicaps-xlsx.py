"""Extract the group's Handicaps spreadsheet into JSON for import-handicaps-xlsx.mjs.

    python3 scripts/extract-handicaps-xlsx.py "Handicaps (5).xlsx" handicaps.json

Kept separate from the importer because the sheet parsing needs openpyxl and the
DB write needs the node `postgres` client. This half is pure and side-effect free.

Sheet shape (verified against the workbook's own summary tabs):
  * `Details` holds one 3-column block per player: Money | Score | Ave.
    Row 5 is the NEWEST round, descending with no gaps.
  * `Score` is strokes over par. `Ave` is derived (mean of the best 6 of the last
    12 scores) and is re-derived on import rather than trusted.
  * Money appears on a subset of the score rows, so the pairing is unambiguous.
  * `Stroking` A2:B20 = current handicap, A25:B43 = Order of Merit. Both derived,
    exported here only as checksums for the importer to validate against.
"""
import json
import sys

import openpyxl


def num(v):
    return v if isinstance(v, (int, float)) and not isinstance(v, bool) else None


def main(src, dest):
    wb = openpyxl.load_workbook(src, data_only=True)
    details, stroking = wb["Details"], wb["Stroking"]

    players = []
    for c in range(1, details.max_column + 1):
        name = details.cell(1, c).value
        if not name:
            continue
        rounds = []
        # Money starts at row 4, scores at row 5; scan from 4 to catch both.
        for r in range(4, details.max_row + 1):
            over = num(details.cell(r, c + 1).value)
            if over is None:
                continue
            rounds.append({"overPar": over, "money": num(details.cell(r, c).value) or 0})
        players.append({"name": str(name).strip(), "rounds": rounds})

    handicaps = {}
    for r in range(2, 21):
        n = stroking.cell(r, 1).value
        if n:
            handicaps[str(n).strip()] = num(stroking.cell(r, 2).value)

    money = {}
    for r in range(25, 44):
        n = stroking.cell(r, 1).value
        if n and str(n).strip():
            money[str(n).strip()] = num(stroking.cell(r, 2).value) or 0

    out = {"source": src, "players": players, "sheetHandicap": handicaps, "sheetMoney": money}
    with open(dest, "w") as f:
        json.dump(out, f, indent=2)

    total = sum(len(p["rounds"]) for p in players)
    print(f"wrote {dest}: {len(players)} players, {total} rounds")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        sys.exit("usage: extract-handicaps-xlsx.py <in.xlsx> <out.json>")
    main(sys.argv[1], sys.argv[2])
